#!/usr/bin/env python
# coding: utf-8

# In[ ]:


import tkinter as tk
from tkinter import ttk
from openpyxl import Workbook, load_workbook
import csv

class StudentRegSys:
    def __init__(self, root):
        self.root = root
        self.root.title("Student Registration System")

        try:
            self.wb = load_workbook("student_data.xlsx")
            self.sheet = self.wb.active
        except FileNotFoundError:
            self.wb = Workbook()
            self.sheet = self.wb.active
            self.sheet.append(["std_Name", "Roll number", "Email_Id", "Course"])

        self.label_name = ttk.Label(root, text="Enter your Name: ")
        self.label_roll_number = ttk.Label(root, text="Enter your Roll Number: ")
        self.label_Email_ID = ttk.Label(root, text="Enter your Email ID:")
        self.label_Course = ttk.Label(root, text="Enter Course Name")

        self.entry_name = ttk.Entry(root)
        self.entry_roll_number = ttk.Entry(root)
        self.entry_Email_ID = ttk.Entry(root)
        self.entry_Course = ttk.Entry(root)

        self.button_add = ttk.Button(root, text="Add Student", command=self.add_student)
        self.button_update = ttk.Button(root, text="Update Student", command=self.update_student)
        self.button_delete = ttk.Button(root, text="Delete Student", command=self.delete_student)
        self.button_display = ttk.Button(root, text="Display Students", command=self.display_students)
        self.button_search = ttk.Button(root, text="Search by Roll Number", command=self.search_student)
        self.button_export = ttk.Button(root, text="Export to CSV", command=self.export_to_csv)

        # Grid layout

        self.label_name.grid(row=0, column=0, padx=10, pady=5, sticky=tk.W)
        self.entry_name.grid(row=0, column=1, padx=10, pady=5, sticky=tk.W)

        self.label_roll_number.grid(row=1, column=0, padx=10, pady=5, sticky=tk.W)
        self.entry_roll_number.grid(row=1, column=1, padx=10, pady=5, sticky=tk.W)

        self.label_Email_ID.grid(row=2, column=0, padx=10, pady=5, sticky=tk.W)
        self.entry_Email_ID.grid(row=2, column=1, padx=10, pady=5, sticky=tk.W)

        self.label_Course.grid(row=3, column=0, padx=10, pady=5, sticky=tk.W)
        self.entry_Course.grid(row=3, column=1, padx=10, pady=5, sticky=tk.W)

        self.button_add.grid(row=4, column=0, pady=10)
        self.button_update.grid(row=4, column=1, pady=10)
        self.button_delete.grid(row=5, column=0, pady=10)
        self.button_display.grid(row=5, column=1, pady=10)
        self.button_search.grid(row=6, column=0, pady=10)
        self.button_export.grid(row=6, column=1, pady=10)

    def add_student(self):
        name = self.entry_name.get()
        roll_number = self.entry_roll_number.get()
        Email_ID = self.entry_Email_ID.get()
        Course = self.entry_Course.get()

        if name and roll_number and Email_ID and Course:
            self.sheet.append([name, roll_number, Email_ID, Course])

            self.wb.save("student_data.xlsx")

            self.entry_name.delete(0, tk.END)
            self.entry_roll_number.delete(0, tk.END)
            self.entry_Email_ID.delete(0, tk.END)
            self.entry_Course.delete(0, tk.END)
        else:
            tk.messagebox.showwarning("Warning", "Please Enter the complete details")

    def update_student(self):
        roll_number = self.entry_roll_number.get()
        selected_item = self.search_student_in_sheet(roll_number)

        if selected_item:
            name = self.entry_name.get()
            Email_ID = self.entry_Email_ID.get()
            Course = self.entry_Course.get()

            selected_index = self.sheet["B"].index(roll_number) + 2  # +2 because Excel is 1-indexed
            self.sheet.cell(row=selected_index, column=1, value=name)
            self.sheet.cell(row=selected_index, column=3, value=Email_ID)
            self.sheet.cell(row=selected_index, column=4, value=Course)

            self.wb.save("student_data.xlsx")

            tk.messagebox.showinfo("Success", "Student details updated successfully.")
        else:
            tk.messagebox.showwarning("Warning", f"No student found with Roll Number {roll_number}.")
       
    def delete_student(self):
        roll_number = self.entry_roll_number.get()
        selected_item = self.search_student_in_sheet(roll_number)

        if selected_item:
            # Delete student data 
            selected_index = self.sheet["B"].index(roll_number) + 2
            self.sheet.delete_rows(selected_index, 1)

            # Save 
            self.wb.save("student_data.xlsx")

            # Clear the entry 
            self.entry_name.delete(0, tk.END)
            self.entry_roll_number.delete(0, tk.END)
            self.entry_Email_ID.delete(0, tk.END)
            self.entry_Course.delete(0, tk.END)

            tk.messagebox.showinfo("Success", "Student deleted successfully.")
        else:
            tk.messagebox.showwarning("Warning", f"No student found with Roll Number {roll_number}.")

    def search_student_in_sheet(self, roll_number):
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            if row[1] == roll_number:
                return row
        return None

    def search_student(self):
        roll_number = self.entry_roll_number.get()
        student_details = self.search_student_in_sheet(roll_number)

        if student_details:
            self.entry_name.delete(0, tk.END)
            self.entry_Email_ID.delete(0, tk.END)
            self.entry_Course.delete(0, tk.END)

            self.entry_name.insert(0, student_details[0])
            self.entry_Email_ID.insert(0, student_details[2])
            self.entry_Course.insert(0, student_details[3])

            tk.messagebox.showinfo("Success", "Student found.")
        else:
            tk.messagebox.showwarning("Warning", f"No student found with Roll Number {roll_number}.")

    def display_students(self):
        display_window = tk.Toplevel(self.root)
        display_window.title("Student information")

        tree = ttk.Treeview(display_window)
        tree["columns"] = ("Name", "Roll Number", "Email ID", 'Course')
        tree.heading("Name", text="Name")
        tree.heading("Roll Number", text="Roll Number")
        tree.heading("Email ID", text="Email ID")
        tree.heading("Course", text="Course")

        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            tree.insert("", "end", values=row)

        tree.pack()

    def export_to_csv(self):
        with open('student_data.csv', 'w', newline='') as file:
            writer = csv.writer(file)
            writer.writerow(["Name", "Roll Number", "Email ID", "Course"])

            for row in self.sheet.iter_rows(min_row=2, values_only=True):
                writer.writerow(row)

        tk.messagebox.showinfo("Success", "Student data exported to student_data.csv.")
        
        

if __name__ == "__main__":
    root = tk.Tk()
    obj = StudentRegSys(root)
    root.mainloop()

   


# In[ ]:





# In[ ]:




